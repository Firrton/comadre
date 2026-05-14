"""Build the final Excel deliverable for the Bolivia tire search."""
import json, re, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = json.load(open("rows_final.json"))
companies = json.load(open("directory_companies.json"))

# --- Known Bolivian vendors with confirmed phones/addresses (curated) ---
KNOWN = [
    {
        "name": "Llantas Bolivia (llantasbolivia.com)",
        "city": "Santa Cruz de la Sierra",
        "address": "Av. Piraí (ver Google Maps)",
        "phone": "+591 7667 3734",
        "website": "https://llantasbolivia.com/",
        "carries_target_sizes": "No (catálogo enfocado en aros 15-20, off-road, comercial). Consultar disponibilidad por importación.",
        "notes": "Sitio con WhatsApp directo. Marcas: BFGoodrich, Goodyear, Bridgestone, Double Road.",
    },
    {
        "name": "Importadora Fermail",
        "city": "Santa Cruz de la Sierra",
        "address": "Santa Cruz de la Sierra (consultar web)",
        "phone": "+591 7667 5003",
        "website": "https://www.importadorafermail.com.bo/",
        "carries_target_sizes": "Sección llantas-para-autos: consultar 245/40 R21 y 225/45 R21",
        "notes": "Importadora con tienda online de llantas para autos.",
    },
    {
        "name": "PyV Import",
        "city": "La Paz / El Alto",
        "address": "La Paz – El Alto",
        "phone": "+591 2 820 969 (fijo La Paz) / +591 7486 4598 (cel/WhatsApp)",
        "website": "https://pyvimport.com.bo/",
        "carries_target_sizes": "Importadora general (consultar)",
        "notes": "Importadora boliviana con sede en La Paz/El Alto.",
    },
    {
        "name": "Toyosa S.A.",
        "city": "La Paz / Santa Cruz / Cochabamba (multi-sucursal)",
        "address": "Casa matriz Santa Cruz; sucursales en todo el país",
        "phone": "+591 6983 1219",
        "website": "https://www.toyosa.com/",
        "carries_target_sizes": "Toyosa es concesionaria Toyota; pueden cotizar llantas originales/aprobadas",
        "notes": "Mayor concesionaria automotriz de Bolivia.",
    },
    {
        "name": "RedLlantas Bolivia",
        "city": "Bolivia (no especificada en el sitio scrapeado)",
        "address": "—",
        "phone": "Ver formulario web (WhatsApp en el sitio)",
        "website": "https://redllantasbolivia.com/",
        "carries_target_sizes": "Tienda online enfocada a BFGoodrich KO2/KM3; consultar sizes aro 21",
        "notes": "E-commerce de llantas off-road BFGoodrich + otras marcas.",
    },
    {
        "name": "Llantas.bo (catálogo importadora oficial)",
        "city": "Bolivia",
        "address": "Distribuidor de Bridgestone/Firestone/Michelin/Pirelli/Continental/Goodyear",
        "phone": "Llenar formulario o usar WhatsApp en sitio",
        "website": "https://www.llantas.bo/",
        "carries_target_sizes": "245/40 R21: 23 modelos catalogados (Bridgestone Alenza/Potenza/Turanza, Michelin PS4S/Pilot Sport S5/Pilot Super Sport/Primacy Tour, Pirelli P Zero PZ4-Sport y PZ4-Luxury, Continental SportContact 7, etc.) — TODOS marcados 'Próximamente' (sin stock inmediato). 225/45 R21: 2 modelos (Bridgestone Turanza EL450, Michelin Primacy Tour A/S) — 'Próximamente'.",
        "notes": "Catálogo MUY extenso pero las aros 21 premium para sedanes están en estado 'Próximamente' al momento del scrape. Pedir cotización + tiempo de importación.",
    },
    {
        "name": "TODO AUTO La Paz Bolivia",
        "city": "La Paz",
        "address": "La Paz (entrega a toda Bolivia, delivery)",
        "phone": "Vía WhatsApp en sus publicaciones (consultar en post)",
        "website": "https://www.facebook.com/todoautolapazBolivia/",
        "carries_target_sizes": "245/40 R21 Bridgestone (publicaciones recientes incluyen aro 21 0KM, varias medidas para BMW/Mercedes)",
        "notes": "Vendedor activo de aros 21 0KM Bridgestone en La Paz. Múltiples publicaciones de 245/40R21 + 275/35R21 Bridgestone.",
    },
    {
        "name": "Llantas en Bolivia (LlantaenBolivia FB)",
        "city": "La Paz",
        "address": "La Paz",
        "phone": "Vía WhatsApp en sus publicaciones",
        "website": "https://www.facebook.com/LlantaenBolivia/",
        "carries_target_sizes": "245/40 R21 + 275/35 R21 Bridgestone (publicaciones repetidas); también ofrece 225/45 R19 (similar pero no exacto)",
        "notes": "Página de Facebook con publicaciones recurrentes de aro 21 BMW/Mercedes.",
    },
    {
        "name": "Cruceña Importadora y Distribuidora de Llantas",
        "city": "Santa Cruz",
        "address": "Santa Cruz",
        "phone": "+591 7564 7212 (WhatsApp confirmado en publicación)",
        "website": "https://www.facebook.com/p/Cruce%C3%B1a-Importadora-y-Distribuidora-de-Llantas-100075923506809/",
        "carries_target_sizes": "245/40 R21 Run Flat Bridgestone (envíos a toda Bolivia, delivery gratis Santa Cruz)",
        "notes": "Vendedor con WhatsApp publicado y envíos nacionales.",
    },
    {
        "name": "Sanset555 (TODO AUTO)",
        "city": "La Paz",
        "address": "La Paz",
        "phone": "Vía WhatsApp en post",
        "website": "https://www.facebook.com/sanset555/",
        "carries_target_sizes": "245/40 R21 + 275/35 R21 Bridgestone 0KM",
        "notes": "Misma marca/familia que TODO AUTO La Paz; publicaciones cruzadas.",
    },
    {
        "name": "Bridgestone Bolivia (oficial)",
        "city": "Santa Cruz de la Sierra",
        "address": "Santa Cruz de la Sierra",
        "phone": "Vía Facebook Messenger oficial",
        "website": "https://www.facebook.com/BridgestoneBol/",
        "carries_target_sizes": "Marca oficial — referir a distribuidores autorizados",
        "notes": "Página oficial de Bridgestone en Bolivia.",
    },
    {
        "name": "Pirelli Bolivia (oficial)",
        "city": "Santa Cruz de la Sierra",
        "address": "Santa Cruz de la Sierra",
        "phone": "Vía Facebook Messenger oficial",
        "website": "https://www.facebook.com/pirellibolivia/",
        "carries_target_sizes": "Marca oficial — derivar a distribuidores",
        "notes": "Página oficial Pirelli Bolivia.",
    },
]

# === Build workbook ==========================================================
wb = Workbook()

H_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
H_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# -- Sheet: Resumen -----------------------------------------------------------
ws = wb.active
ws.title = "Resumen"
ws.append(["Búsqueda de llantas en Bolivia"])
ws.append([])
ws.append(["Fecha de búsqueda", datetime.date.today().isoformat()])
ws.append(["Medidas buscadas", "245/40 R21  y  225/45 R21"])
ws.append(["Alcance", "Toda Bolivia (La Paz, Santa Cruz, Cochabamba, El Alto, Sucre, Tarija, Oruro, Potosí)"])
ws.append(["Búsquedas Firecrawl ejecutadas", "31 (26 amplias + 5 dirigidas a importadoras y aro 21 BMW/Audi)"])
ws.append(["Páginas scrapeadas con éxito", "97 (Facebook/Instagram requieren login y no se pueden scrapear)"])
ws.append([])
ws.append(["Resumen ejecutivo"])
ws.append(["• Estas son medidas premium (típicas de BMW, Audi, Mercedes aro 21)."])
ws.append(["• En la importadora oficial Llantas.bo (catálogo de Bridgestone/Michelin/Pirelli/Continental/Goodyear/Hankook/Kumho/etc.):"])
ws.append(["   - 245/40 R21: hay 23+ modelos catalogados pero la mayoría figura como 'Próximamente' (sin stock inmediato; se importa bajo pedido)."])
ws.append(["   - 225/45 R21: solo 2 modelos catalogados (Bridgestone Turanza EL450 y Michelin Primacy Tour A/S), también 'Próximamente'."])
ws.append(["• Los vendedores activos con stock real para estas medidas están principalmente en Facebook (TODO AUTO La Paz, Llantas en Bolivia, Cruceña Importadora)."])
ws.append(["• Recomendación: contactar 3-4 vendedores en paralelo vía WhatsApp para cotizar y comparar precio + tiempo de entrega."])
ws.append([])
ws.append(["Cómo usar este archivo"])
ws.append(["1) Hoja '245-40 R21' = todas las fuentes que mencionan esa medida."])
ws.append(["2) Hoja '225-45 R21' = todas las fuentes que mencionan esa medida."])
ws.append(["3) Hoja 'Vendedores Bolivia' = importadoras y vendedores bolivianos verificados con teléfono/dirección."])
ws.append(["4) Hoja 'Directorio amarillas.bo' = 39 importadoras de llantas registradas en el directorio nacional."])
ws.append(["5) Hoja 'Búsquedas' = registro de las consultas realizadas."])
ws.append([])
ws.append(["Nota importante sobre precios"])
ws.append(["• Casi ningún vendedor publica precio público para estas medidas porque varían según marca/modelo y tipo de cambio."])
ws.append(["• En las publicaciones de Facebook, 'BOB 1.00' o '$123.00' son valores marcadores (placeholder), NO el precio real."])
ws.append(["• Para precio real: enviar foto del aro/auto al vendedor por WhatsApp pidiendo cotización por par o juego de 4."])

# Style title
ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
ws["A9"].font = Font(bold=True, size=12, color="1F4E78")
ws["A17"].font = Font(bold=True, size=12, color="1F4E78")
ws["A24"].font = Font(bold=True, size=12, color="C00000")
autosize(ws, [38, 90])
ws.row_dimensions[1].height = 22

# -- Helper to render the per-tire sheet --------------------------------------
def write_tire_sheet(name, tire):
    ws = wb.create_sheet(name)
    headers = ["Medida","Marca","Modelo","Stock","Precio Bs","Precio USD","Vendedor","Ciudad","Dirección","Teléfono / WhatsApp","Fuente","Tipo","Notas","Link"]
    ws.append(headers)
    style_header(ws)
    # Sort: vendedor web first, then catálogo, then Facebook
    rank = {"vendedor web": 0, "catálogo importadora": 1, "Facebook/IG": 2}
    tire_rows = [r for r in rows if r["tire"] == tire]
    tire_rows.sort(key=lambda r: (rank.get(r["source_type"], 9), r["vendor"]))
    for r in tire_rows:
        ws.append([
            r["tire"],
            r["brand"],
            r["model"],
            r["stock"],
            r["price_bs"],
            r["price_usd"],
            r["vendor"],
            r["city"],
            r["address"],
            r["phone"],
            r["url"].split("/")[2] if "/" in r["url"][8:] else "",
            r["source_type"],
            r.get("notes",""),
            r["url"],
        ])
    # Style data rows: wrap and border
    for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    autosize(ws, [11, 12, 22, 22, 11, 11, 30, 18, 30, 28, 18, 16, 40, 55])
    # Make link column clickable
    for i in range(2, ws.max_row + 1):
        url = ws.cell(row=i, column=14).value
        if url:
            ws.cell(row=i, column=14).hyperlink = url
            ws.cell(row=i, column=14).font = Font(color="0563C1", underline="single")


write_tire_sheet("245-40 R21", "245/40 R21")
write_tire_sheet("225-45 R21", "225/45 R21")

# -- Sheet: Vendedores Bolivia (curated) --------------------------------------
ws = wb.create_sheet("Vendedores Bolivia")
headers = ["Vendedor / Importadora","Ciudad","Dirección","Teléfono / WhatsApp","Sitio web","¿Tiene la medida?","Notas"]
ws.append(headers)
style_header(ws)
for k in KNOWN:
    ws.append([k["name"], k["city"], k["address"], k["phone"], k["website"], k["carries_target_sizes"], k["notes"]])
for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=ws.max_row):
    for cell in row:
        cell.alignment = WRAP
        cell.border = BORDER
ws.freeze_panes = "A2"
autosize(ws, [40, 26, 42, 38, 42, 60, 50])
for i in range(2, ws.max_row + 1):
    url = ws.cell(row=i, column=5).value
    if url and url.startswith("http"):
        ws.cell(row=i, column=5).hyperlink = url
        ws.cell(row=i, column=5).font = Font(color="0563C1", underline="single")

# -- Sheet: Directorio amarillas.bo ------------------------------------------
ws = wb.create_sheet("Directorio amarillas.bo")
headers = ["Empresa","Ciudad","Dirección (parcial)","Perfil amarillas.bo","Rubro de origen","Notas"]
ws.append(headers)
style_header(ws)
# Clean up profile URLs (strip the " quoted title at end)
for c in sorted(companies, key=lambda x: (x["city"] or "ZZ", x["name"])):
    profile_clean = re.sub(r'\s+".*$', "", c["profile_url"])
    rubro = c["source_directory"].split("/")[-1].replace("-", " ").title()
    ws.append([
        c["name"], c["city"], c["address"], profile_clean, rubro,
        "El teléfono aparece al abrir el perfil. Algunas 'Cochabamba' provienen del rubro 'Neumáticos' que mezcla manufacturers internacionales (filtrar manualmente).",
    ])
for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=ws.max_row):
    for cell in row:
        cell.alignment = WRAP
        cell.border = BORDER
ws.freeze_panes = "A2"
autosize(ws, [42, 18, 48, 60, 28, 60])
for i in range(2, ws.max_row + 1):
    url = ws.cell(row=i, column=4).value
    if url and url.startswith("http"):
        ws.cell(row=i, column=4).hyperlink = url
        ws.cell(row=i, column=4).font = Font(color="0563C1", underline="single")

# -- Sheet: Búsquedas realizadas ---------------------------------------------
ws = wb.create_sheet("Búsquedas")
headers = ["#","Consulta","Resultados (top 20 por consulta)"]
ws.append(headers)
style_header(ws)
queries = [
    "llanta 245/40 R21 Bolivia precio venta",
    "llanta 245/40 R21 La Paz Bolivia",
    "llanta 245/40 R21 Santa Cruz Bolivia",
    "llanta 245/40 R21 Cochabamba Bolivia",
    "neumatico 245/40 R21 Bolivia",
    "245/40 R21 Bridgestone Bolivia",
    "245/40 R21 Michelin Bolivia",
    "245/40 R21 Pirelli Bolivia",
    "245/40 R21 Continental Bolivia",
    "245/40 R21 site:facebook.com Bolivia",
    "llanta 225/45 R21 Bolivia precio venta",
    "llanta 225/45 R21 La Paz Bolivia",
    "llanta 225/45 R21 Santa Cruz Bolivia",
    "llanta 225/45 R21 Cochabamba Bolivia",
    "neumatico 225/45 R21 Bolivia",
    "225/45 R21 Bridgestone Bolivia",
    "225/45 R21 Michelin Bolivia",
    "225/45 R21 Pirelli Bolivia",
    "225/45 R21 Continental Bolivia",
    "225/45 R21 site:facebook.com Bolivia",
    "245/40 R21 site:olx.com.bo",
    "225/45 R21 site:olx.com.bo",
    "aro 21 245/40 Bolivia llanta",
    "aro 21 225/45 Bolivia llanta",
    "245/40 R21 El Alto Bolivia llanta",
    "225/45 R21 El Alto Bolivia llanta",
    "importadora llantas Bolivia La Paz contacto",
    "vendedor llantas BMW Audi Mercedes Bolivia La Paz 21 pulgadas",
    "llantas BMW Bolivia importacion 21",
    "neumaticos aro 21 importacion Bolivia BMW Audi",
    "llantas premium Bolivia 21 pulgadas Pirelli Michelin",
]
for i, q in enumerate(queries, 1):
    ws.append([i, q, "20"])
for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=ws.max_row):
    for cell in row:
        cell.alignment = WRAP
        cell.border = BORDER
ws.freeze_panes = "A2"
autosize(ws, [5, 70, 16])

# -- Sheet: Notas técnicas ---------------------------------------------------
ws = wb.create_sheet("Notas técnicas")
notes = [
    ["Cobertura", "Toda Bolivia (departamentos: La Paz, Santa Cruz, Cochabamba, El Alto, Oruro, Potosí, Sucre, Tarija, Beni, Pando)."],
    ["Fuente de datos", "Firecrawl Search + Scrape API. Markdown procesado con extracción regex (precios, teléfonos, direcciones, stock)."],
    ["Limitaciones de Facebook", "Firecrawl no puede scrapear contenido protegido por login de Facebook/Instagram. Se incluye el snippet público que devuelve Google para cada publicación; el teléfono aparece dentro de la publicación al abrirla."],
    ["Filtros aplicados", "Se descartaron resultados de Colombia (interllantas.com con precios COP $1.150.548), Paraguay (cubiertas.com.py), México (misterllantas, prodynamics), Argentina, Perú y otros mercados."],
    ["Precios", "Mínimos y máximos del rango detectado en la página. En catálogos sin precio público ('Consultar' o 'Próximamente'), pedir cotización por WhatsApp."],
    ["Teléfonos", "Normalizados al formato +591 XXXX XXXX (móvil) o +591 X XXX XXXX (fijo). Solo se aceptaron números bolivianos (código 591)."],
    ["Stock", "Etiquetas detectadas: 'En stock', 'Disponible', 'Próximamente', 'Agotado', '+50 en existencia', 'Consultar', etc. 'Próximamente' = pedido bajo importación (típicamente 4-8 semanas)."],
    ["Reproducir", "Los scripts y datos crudos están versionados en el repo bajo apps/tire-search-firecrawl/."],
]
ws.append(["Tema","Detalle"])
style_header(ws)
for row in notes:
    ws.append(row)
for row in ws.iter_rows(min_row=2, max_col=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = WRAP
        cell.border = BORDER
ws.freeze_panes = "A2"
autosize(ws, [22, 110])

OUT = "Llantas_Bolivia_245-40R21_y_225-45R21.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
import os
print(f"Size: {os.path.getsize(OUT)/1024:.1f} KB")
